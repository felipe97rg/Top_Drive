import pandas as pd
import os
import math
import io
import traceback 

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment 
from PIL import Image as PILImage 

# --- 1. CONFIGURACIÓN ---
RUTA_EXCEL_ENTRADA = r"\\192.168.1.2\cenyt-proyectos\CEN-223_TD GUAFITA\2.INFOENTRADA\DATOS_FINALES\SIU_consolidado.xlsx"
CARPETA_FOTOS = r"\\192.168.1.2\cenyt-proyectos\CEN-223_TD GUAFITA\2.INFOENTRADA\DATOS_FINALES\FOTOS_TRATADAS"
RUTA_EXCEL_SALIDA = r"\\192.168.1.2\cenyt-proyectos\CEN-223_TD GUAFITA\2.INFOENTRADA\DATOS_FINALES\Reporte_Fotografico_1.xlsx"

# --- 2. CONSTANTES DE FORMATO ---
MAX_FILAS_POR_HOJA = 30
CELL_HEIGHT_PTS = 80 
CELL_WIDTH_PTS = 20 
IMG_MAX_WIDTH_PX = 145 
IMG_MAX_HEIGHT_PX = 105

# Columnas de observaciones generales
COLUMNAS_OBSERVACIONES = [
    "Configuracion",'Configuracion_Observaciones',
    "Disposicion", "Disposicion_Observaciones",'Circuito_Observaciones', 'Observaciones', 'Terreno_Observaciones', 
    'Apoyo_Observacion', 'Disposicion_Observaciones',
    'Observacion_Cruceta', 'Aisladores_Observaciones', 'DPS_Observaciones',
    'Observaciones_Equipos', 'Afloramiento_Observaciones', 'SPT_Observaciones',
    'Otras_Observaciones'
]
# Columnas de condición del Apoyo (Sí/No)
COLUMNAS_CONDICION = [
    'Apoyo_BuenEstado', 'Apoyo_Averias', 'Apoyo_Porosidad', 
    'Apoyo_Fractura', 'Apoyo_Oxido', 'Apoyo_Humedad', 'Apoyo_Vandalizado'
]

# Columnas de condición de la Cruceta (Sí/No)
COLUMNAS_CRUCETA = [
    'Cruceta_BuenEstado',
    'Cruceta_Oxido',
    'Cruceta_Averias',
    'Cruceta_Fractura',
    'Cruceta_Humedad'
]

# Columnas finales a mostrar
COLUMNAS_FINALES = [
    'Circuito', 'Estructura_Tag', 'Latitud_Manual', 'Longitud_Manual',
    'Observaciones Generales', 'Foto_1', 'Foto_2'
]

# --- 3. NUEVA FUNCIÓN DE CORRECCIÓN DE TILDES ---

def corregir_tildes(texto):
    """
    Intenta corregir el problema de 'mojibake' (e.g., "MetÃ¡licos" -> "Metálicos").
    Este es un problema común cuando texto UTF-8 se interpreta erróneamente como latin-1.
    """
    # Si es nulo (NaN, NaT, None), lo devolvemos tal cual
    if pd.isna(texto):
        return texto
    
    # Convertimos a string por si acaso (ej. si es un número)
    texto_str = str(texto)
    
    try:
        # El truco: codificar en latin-1 (que trata cada byte como un caracter)
        # y luego decodificar como UTF-8 (que interpreta la secuencia de bytes)
        return texto_str.encode('latin-1').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        # Si falla (p.ej., ya estaba bien, es un número, o es otro tipo de error),
        # devolvemos el texto original (convertido a string).
        return texto_str

# --- 4. FUNCIÓN AUXILIAR PARA CONCATENAR (MODIFICADA) ---

def crear_texto_observaciones(row):
    """
    Crea el texto en formato "diccionario" para las observaciones.
    Añade Tipo, Condición Apoyo (sin 'Apoyo_') y Condición Cruceta (sin 'Cruceta_').
    ¡AHORA CORRIGE TILDES!
    """
    partes = [] 
    
    # --- 1. Lógica (Tipo y Subtipo) ---
    tipo_partes = []
    # APLICAMOS CORRECCIÓN
    val_tipo = corregir_tildes(row.get('Apoyo_Tipo'))
    val_subtipo = corregir_tildes(row.get('Apoyo_Subtipo'))
    
    if pd.notna(val_tipo) and str(val_tipo).strip() != "":
        tipo_partes.append(str(val_tipo).strip())
    if pd.notna(val_subtipo) and str(val_subtipo).strip() != "":
        tipo_partes.append(str(val_subtipo).strip())
        
    if tipo_partes:
        partes.append(f"Tipo Estructura: {' '.join(tipo_partes)}")

    # --- 2. Lógica (Condiciones "Sí" - Apoyo) ---
    condicion_partes = []
    for col in COLUMNAS_CONDICION:
        if col in row.index:
            valor = row[col]
            if pd.notna(valor) and str(valor).strip().lower() == 'sí':
                nombre_limpio = col.replace('Apoyo_', '')
                condicion_partes.append(nombre_limpio)
    
    if condicion_partes:
        partes.append(f"Condición Apoyo: {', '.join(condicion_partes)}")

    # --- 3. Lógica (Condiciones "Sí" - Cruceta) ---
    cruceta_partes = []
    for col in COLUMNAS_CRUCETA: 
        if col in row.index:
            valor = row[col]
            if pd.notna(valor) and str(valor).strip().lower() == 'sí':
                nombre_limpio = col.replace('Cruceta_', '')
                cruceta_partes.append(nombre_limpio)
    
    if cruceta_partes:
        partes.append(f"Condición Cruceta: {', '.join(cruceta_partes)}")

    # --- 4. Lógica (Observaciones generales tipo diccionario) ---
    for col in COLUMNAS_OBSERVACIONES:
        if col in row.index:
            # APLICAMOS CORRECCIÓN
            valor = corregir_tildes(row[col])
            
            if pd.notna(valor) and str(valor).strip() != "":
                partes.append(f"{col}: {valor}")

    # --- 5. Final ---
    return ", ".join(partes)

# --- 5. FUNCIÓN AUXILIAR PARA INSERTAR IMÁGENES (Sin Cambios) ---

def insertar_imagen_en_celda(ws, cell, nombre_foto, carpeta_base):
    if pd.isna(nombre_foto):
        cell.value = "N/A"
        return

    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ruta_foto = os.path.join(carpeta_base, nombre_foto)

    try:
        img = PILImage.open(ruta_foto)
        img.thumbnail((IMG_MAX_WIDTH_PX, IMG_MAX_HEIGHT_PX))
        
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        
        oxl_img = OpenpyxlImage(img_buffer)
        oxl_img.anchor = cell.coordinate
        ws.add_image(oxl_img)

    except FileNotFoundError:
        print(f"  Advertencia: No se encontró la foto '{nombre_foto}'")
        cell.value = "Foto no encontrada"
    except Exception as e:
        print(f"  Error al procesar la foto '{nombre_foto}': {e}")
        cell.value = "Error al cargar"


# --- 6. SCRIPT PRINCIPAL (Sin cambios en la lógica) ---

print("Iniciando proceso de reporte fotográfico...")
writer = None 

try:
    print(f"Leyendo {RUTA_EXCEL_ENTRADA}...")
    df_base = pd.read_excel(RUTA_EXCEL_ENTRADA)

    print("Concatenando observaciones...")
    # Esta línea ahora usará la función MODIFICADA
    df_base['Observaciones Generales'] = df_base.apply(crear_texto_observaciones, axis=1)

    if 'Foto_1' not in df_base.columns: df_base['Foto_1'] = pd.NA
    if 'Foto_2' not in df_base.columns: df_base['Foto_2'] = pd.NA
        
    df_final = df_base[COLUMNAS_FINALES]

    # --- ¡CORRECCIÓN AÑADIDA AQUÍ! ---
    TEXTO_URL_A_QUITAR = "https://github.com/felipe97rg/Top_Drive/tree/main/DATOS_FINALES/FOTOS_TRATADAS/"
    print(f"Limpiando prefijo de URL en columnas de fotos...")
    df_final['Foto_1'] = df_final['Foto_1'].str.replace(TEXTO_URL_A_QUITAR, '', regex=False)
    df_final['Foto_2'] = df_final['Foto_2'].str.replace(TEXTO_URL_A_QUITAR, '', regex=False)
    # --- FIN DE LA CORRECCIÓN ---

    writer = pd.ExcelWriter(RUTA_EXCEL_SALIDA, engine='openpyxl')
    
    grupos_circuitos = df_final.groupby('Circuito')

    for nombre_circuito, df_circuito in grupos_circuitos:
        print(f"\nProcesando Circuito: {nombre_circuito}")
        
        num_filas = len(df_circuito)
        num_hojas = math.ceil(num_filas / MAX_FILAS_POR_HOJA)
        
        if num_hojas == 0:
            continue

        for i in range(num_hojas):
            if num_hojas > 1:
                nombre_hoja = f"{nombre_circuito}_Pt{i+1}"
            else:
                nombre_hoja = str(nombre_circuito)
            
            nombre_hoja = nombre_hoja[:31]
            print(f"  Creando hoja: {nombre_hoja}")
            
            inicio = i * MAX_FILAS_POR_HOJA
            fin = (i + 1) * MAX_FILAS_POR_HOJA
            df_pagina = df_circuito.iloc[inicio:fin]
            
            df_pagina = df_pagina.copy()
            
            df_pagina.to_excel(writer, sheet_name=nombre_hoja, index=False)
            
            ws = writer.sheets[nombre_hoja]
            
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 50 
            ws.column_dimensions['F'].width = CELL_WIDTH_PTS
            ws.column_dimensions['G'].width = CELL_WIDTH_PTS

            for idx, row_num_excel in enumerate(range(2, len(df_pagina) + 2)):
                
                ws.row_dimensions[row_num_excel].height = CELL_HEIGHT_PTS
                
                nombre_foto1 = df_pagina.iloc[idx]['Foto_1']
                nombre_foto2 = df_pagina.iloc[idx]['Foto_2']
                
                cell_foto1 = ws[f'F{row_num_excel}']
                cell_foto2 = ws[f'G{row_num_excel}']
                
                insertar_imagen_en_celda(ws, cell_foto1, nombre_foto1, CARPETA_FOTOS)
                insertar_imagen_en_celda(ws, cell_foto2, nombre_foto2, CARPETA_FOTOS)

    print("\nProceso casi completo. Guardando archivo...")

except Exception as e:
    print(f"\n¡ERROR! Ocurrió un problema: {e}")
    traceback.print_exc()

finally:
    if writer:
        writer.close()
        print(f"¡Éxito! Archivo guardado en: {RUTA_EXCEL_SALIDA}")